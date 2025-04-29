import time
import traceback
import pandas as pd
from io import StringIO
import csv
#追加
import os


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
#追加
from datetime import datetime, timedelta

# 【設定項目】Excel ファイルのパス
EXCEL_FILE = "Enhance_広告枠検索対象.xlsx"
SHEET_NAME = "日次レポート"

# 【設定項目】ログイン情報
LOGIN_URL = "https://admin.fam-8.net/report/index.php"
ID = "admin"
PASSWORD = "fhC7UPJiforgKTJ8"

# 【設定項目】出力 CSV
CSV_FILE = "enhance_utf8(日次レポート).csv"

def get_all_adframe_ids():
    """Excel の日次レポート・日次レポート (マイナビ)の A3 から空行までの広告枠IDを取得"""
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)

        def fetch_ids(sheet_name):
            """指定したシートの A3 から空行までの広告枠IDを取得"""
            ws = wb[sheet_name]
            adframe_ids = []
            for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
                if row[0] is None:
                    break
                adframe_ids.append(str(row[0]).strip())

            # 【デバッグ】取得した広告枠IDを出力
            print(f"[DEBUG] {sheet_name} の取得広告枠ID数: {len(adframe_ids)}")
            print(f"[DEBUG] {sheet_name} の取得広告枠ID: {adframe_ids[:5]}...")  # 最初の5件だけ表示

            return adframe_ids

        # 2つのシートからそれぞれ取得
        adframe_ids_report = fetch_ids("日次レポート")
        adframe_ids_mynavi = fetch_ids("日次レポート (マイナビ)")

        wb.close()

        return adframe_ids_report, adframe_ids_mynavi

    except Exception as e:
        print("[ERROR] Excel の読み込みでエラー発生")
        print(traceback.format_exc())
        return [], []




def setup_driver():
    """ WebDriver をセットアップして返す """
    print("[INFO] WebDriver を起動します")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    driver.get(LOGIN_URL)
    print("[INFO] WebDriver の起動完了")
    return driver

def login(driver):
    """ ログイン処理 """
    try:
        print("[INFO] ログイン処理を開始します")
        wait = WebDriverWait(driver, 10)
        driver.find_element(By.XPATH, '//*[@id="topmenu"]/tbody/tr[2]/td/div[1]/form/div/table/tbody/tr[1]/td/input').send_keys(ID)
        driver.find_element(By.XPATH, '//*[@id="topmenu"]/tbody/tr[2]/td/div[1]/form/div/table/tbody/tr[2]/td/input').send_keys(PASSWORD)
        driver.find_element(By.XPATH, '//*[@id="topmenu"]/tbody/tr[2]/td/div[1]/form/div/table/tbody/tr[3]/td/input[2]').click()
        time.sleep(2)
        print("[INFO] ログイン完了")
    except Exception as e:
        print("[ERROR] ログイン処理でエラー発生")
        print(traceback.format_exc())
        exit(1)

#サイト側が重いので明示的に待機を追記
def search_adframe(driver, adframe_id, first_search=True):
    """広告枠ID を検索する（1回目と2回目以降で処理を分岐）"""
    try:
        print(f"[INFO] {'初回検索' if first_search else '再検索'}: 広告枠ID {adframe_id} を検索")

        wait = WebDriverWait(driver, 10)

        if first_search:
            # 1回目のみ実行する処理
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="sidemenu"]/div[3]/a[4]/div'))).click()
            time.sleep(1)  # **クリック後の反映待機**
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="display_modesummary_mode"]'))).click()
            time.sleep(0.5)
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="main_area"]/form/div[1]/select[2]/option[2]'))).click()
            time.sleep(0.5)
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="main_area"]/form/div[1]/select[3]/option[1]'))).click()
            time.sleep(0.5)
            wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="main_area"]/form/div[1]/table[2]/tbody/tr[1]/td/select[3]/option[1]'))).click()
            time.sleep(0.5)

        # **検索窓のクリア（全回実行）**
        input_element = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="main_area"]/form/div[1]/input[7]')))
        input_element.clear()
        input_element.send_keys(adframe_id)
        time.sleep(0.5)  # **入力後の安定待機**

        # **検索ボタン押下（全回実行）**
        search_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="main_area"]/form/div[1]/input[10]')))
        search_button.click()
        time.sleep(1.5)  # **検索結果が安定するまで待機**

        # **検索結果の読み込みを待機（最大10秒）**
        retries = 3
        for attempt in range(retries):
            try:
                wait.until(EC.presence_of_element_located((By.ID, "tbl_data")))
                break  # **成功したらループを抜ける**
            except StaleElementReferenceException:
                if attempt < retries - 1:
                    print(f"[WARNING] StaleElementReferenceException 発生。再試行 ({attempt+1}/{retries})")
                    time.sleep(1)
                else:
                    raise  # **最終リトライでも失敗したらエラーを投げる**

        print(f"[INFO] 広告枠ID {adframe_id} の検索完了")

    except Exception:
        print(f"[ERROR] 広告枠ID {adframe_id} の検索でエラー発生")
        print(traceback.format_exc())




def extract_table_data(driver, target_adframe_id):
    """ 検索結果の表データを取得し、必要なカラムのみ抽出 """
    try:
        print(f"[INFO] 広告枠ID {target_adframe_id} の検索結果を取得")

        wait = WebDriverWait(driver, 10)

        # 検索結果テーブルの取得をリトライ（最大3回）
        retries = 3
        table_html = None
        for attempt in range(retries):
            try:
                table_element = wait.until(EC.presence_of_element_located((By.ID, "tbl_data")))
                table_html = table_element.get_attribute("outerHTML")
                break  # 成功したらループ終了
            except selenium.common.exceptions.StaleElementReferenceException:
                print(f"[WARNING] 検索結果が更新されたため、再取得を試みます ({attempt + 1}/{retries})")
                time.sleep(1)  # 1秒待機してリトライ

        # データが取得できなかった場合は終了
        if not table_html:
            print(f"[ERROR] 広告枠ID {target_adframe_id} のデータ取得に失敗しました")
            return None

        # Pandas の read_html でテーブルデータを解析
        df_list = pd.read_html(StringIO(table_html), header=0)

        if not df_list or len(df_list) == 0:
            print(f"[ERROR] 広告枠ID {target_adframe_id} のデータテーブルが取得できませんでした")
            return None

        df = df_list[0]

        # 【デバッグ】取得したデータのカラム名を表示
        print("[DEBUG] 取得したデータのカラム名:", df.columns.tolist())

        # NaN の処理
        df = df.fillna(0)

        # 広告枠ID のデータ型を統一
        df["広告枠ID"] = df["広告枠ID"].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

        # 検索対象の広告枠IDと一致する行のみ取得
        filtered_df = df[df["広告枠ID"] == str(target_adframe_id)]

        if filtered_df.empty:
            print(f"[WARNING] 広告枠ID {target_adframe_id} に該当するデータが見つかりませんでした")
            return None

        # 必要なカラムのみ抽出
        columns_to_keep = ["広告枠ID", "広告枠名", "Imp", "Click", "ネット"]
        filtered_df = filtered_df.loc[:, columns_to_keep]

        print("[DEBUG] 抽出したデータ:")
        print(filtered_df)

        return filtered_df

    except Exception as e:
        print("[ERROR] 表データの取得でエラー発生")
        print(traceback.format_exc())
        return None




        df = df_list[0]

        # 【デバッグ】取得データのカラム名を表示
        print("[DEBUG] 取得したデータのカラム名:", df.columns.tolist())

        # NaN の処理
        df = df.fillna(0)

        # 広告枠ID のデータ型を統一（str に変換して一致判定を確実に）
        df["広告枠ID"] = df["広告枠ID"].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

        # 検索対象の広告枠IDと一致する行のみ取得
        filtered_df = df[df["広告枠ID"] == str(target_adframe_id)]

        if filtered_df.empty:
            print(f"[WARNING] 広告枠ID {target_adframe_id} に該当するデータが見つかりませんでした")
            return None

        # 必要なカラムのみ抽出（カラム位置を修正）
        columns_to_keep = {
            "広告枠ID": 0,
            "広告枠名": 1,
            "Imp": 8,      # 実際の列番号を確認して修正すること
            "Click": 9,    # 実際の列番号を確認して修正すること
            "ネット": 15    # 実際の列番号を確認して修正すること
        }

        filtered_df = filtered_df.iloc[:, list(columns_to_keep.values())]
        filtered_df.columns = list(columns_to_keep.keys())  # カラム名を統一

        print("[DEBUG] 抽出したデータ:")
        print(filtered_df)

        return filtered_df if not filtered_df.empty else None


        # NaN の処理
        df = df.fillna(0)  # NaN を 0 に変換

        # 広告枠ID のデータ型を統一（str に変換して一致判定を確実に）
        df["広告枠ID"] = df["広告枠ID"].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

        # 検索対象の広告枠IDと一致する行のみ取得
        filtered_df = df[df["広告枠ID"] == str(target_adframe_id)]

        # 必要なカラムのみ抽出
        columns_to_keep = ["広告枠ID", "広告枠名", "Imp", "Click", "ネット"]
        filtered_df = filtered_df.loc[:, columns_to_keep]

        print("[DEBUG] 抽出したデータ:")
        print(filtered_df)

        return filtered_df

    except Exception as e:
        print("[ERROR] 表データの取得でエラー発生")
        print(traceback.format_exc())
        return None

def save_to_csv(data, csv_file):
    """ データを CSV に保存（追記対応） """
    if data is None or data.empty:
        print(f"[WARNING] {csv_file} に保存するデータがありません")
        return

    # 【デバッグ】保存前のデータを表示
    print(f"[DEBUG] {csv_file} に保存するデータのカラム: {list(data.columns)}")
    print(f"[DEBUG] {csv_file} に保存するデータの先頭5行:")
    print(data.head())

    # 期待するカラム順を定義
    expected_columns = ["広告枠ID", "広告枠名", "Imp", "Click", "ネット"]
    actual_columns = list(data.columns)

    # カラム順のチェック
    if actual_columns != expected_columns:
        print(f"[ERROR] {csv_file} のデータカラムが不正です。")
        print(f"想定: {expected_columns}")
        print(f"実際: {actual_columns}")
        return

    # 処理実行日の前日を取得
    previous_day = (datetime.today() - timedelta(days=1)).strftime("%Y/%m/%d")

    # 日付列を追加
    data.insert(0, "日付", previous_day)

    # **Windows環境で使えない `:` を `_` に置換**
    csv_file = csv_file.replace(":", "_")

    # `.csv` がついていない場合に追加
    if not csv_file.endswith(".csv"):
        csv_file += ".csv"

    # 既存の CSV があるか確認（1回目は新規作成、2回目以降は追記）
    file_exists = os.path.exists(csv_file)

    try:
        # CSV に保存（追記モードで追加）
        data.to_csv(csv_file, mode='a', header=not file_exists, index=False, encoding="utf-8-sig")

        print(f"[INFO] CSV ファイルに保存完了（追記）: {csv_file}")

        # 【デバッグ】保存後のファイルサイズ確認
        print(f"[DEBUG] {csv_file} のファイルサイズ: {os.path.getsize(csv_file)} バイト")

    except Exception as e:
        print(f"[ERROR] {csv_file} への書き込み中にエラー発生")
        print(traceback.format_exc())




if __name__ == "__main__":
    adframe_ids_report, adframe_ids_mynavi = get_all_adframe_ids()

    if not adframe_ids_report and not adframe_ids_mynavi:
        print("[ERROR] 取得できる広告枠IDがありません。処理を終了します。")
        exit(1)

    driver = setup_driver()
    login(driver)

    # 【日次レポート】の処理
    for i, adframe_id in enumerate(adframe_ids_report):
        first_search = (i == 0)  # 最初の検索だけ True
        search_adframe(driver, adframe_id, first_search)  # 検索実行
        data = extract_table_data(driver, adframe_id)
        if data is not None and not data.empty:
            save_to_csv(data, "enhance_utf8(日次レポート).csv")  # 日次レポート用

    # 【日次レポート (マイナビ)】の処理
    for i, adframe_id in enumerate(adframe_ids_mynavi):
        first_search = (i == 0)  # 最初の検索だけ True
        search_adframe(driver, adframe_id, first_search)  # 検索実行
        data = extract_table_data(driver, adframe_id)
        if data is not None and not data.empty:
            save_to_csv(data, "enhance_utf8(日次レポート_マイナビ).csv")  # マイナビ用

    print("[INFO] すべての広告枠IDの検索が完了")
    driver.quit()